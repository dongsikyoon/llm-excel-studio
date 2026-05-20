from __future__ import annotations

import ast
import importlib
import io
import signal
import threading
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

from core.files.manager import RESULT_DIR, list_files, read_file

# pandas/numpy는 이미 pd/np로 제공 — import 시 자동 제거
_PROVIDED = frozenset({"pandas", "numpy"})

# 허용되는 표준 라이브러리
_ALLOWED_MODULES = frozenset({
    "re", "math", "json", "datetime", "collections",
    "itertools", "functools", "string", "decimal",
    "statistics", "copy", "random", "hashlib",
})

# 차단 모듈 (보안)
_BLOCKED_MODULES = frozenset({
    "os", "subprocess", "sys", "shutil",
    "socket", "http", "urllib", "requests", "httpx",
    "pathlib", "glob", "pickle", "shelve", "marshal",
    "ctypes", "multiprocessing", "threading",
    "signal", "atexit", "importlib", "code",
})

_BLOCKED_BUILTINS = frozenset({
    "exec", "eval", "compile",
    "open", "input", "breakpoint",
    "globals", "locals", "vars",
    "getattr", "setattr", "delattr", "memoryview",
})


@dataclass
class ExecutionResult:
    success: bool
    output: str = ""
    error: str = ""
    result_df: pd.DataFrame | None = None
    saved_files: list[str] = field(default_factory=list)


def _strip_provided_imports(code: str) -> str:
    """pandas/numpy import 문을 제거 — 이미 pd/np로 제공됨."""
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return code

    new_body = []
    for node in tree.body:
        if isinstance(node, ast.Import):
            kept = [a for a in node.names if a.name.split(".")[0] not in _PROVIDED]
            if kept:
                node.names = kept
                new_body.append(node)
        elif isinstance(node, ast.ImportFrom):
            if node.module and node.module.split(".")[0] not in _PROVIDED:
                new_body.append(node)
        else:
            new_body.append(node)

    tree.body = new_body
    return ast.unparse(tree)


def _validate(code: str) -> list[str]:
    try:
        tree = ast.parse(code)
    except SyntaxError as e:
        return [f"문법 오류: {e}"]

    issues: list[str] = []
    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            names = (
                [a.name for a in node.names]
                if isinstance(node, ast.Import)
                else ([node.module] if node.module else [])
            )
            for name in names:
                root = name.split(".")[0]
                if root in _BLOCKED_MODULES:
                    issues.append(f"보안상 허용되지 않는 모듈: {name}")
                elif root not in _ALLOWED_MODULES:
                    issues.append(f"허용되지 않는 모듈: {name} (허용: {', '.join(sorted(_ALLOWED_MODULES))})")
        elif isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name) and node.func.id in _BLOCKED_BUILTINS:
                issues.append(f"허용되지 않는 함수: {node.func.id}")
    return issues


def _make_safe_import():
    # numpy·pandas는 np/pd로 이미 제공되므로 내부 서브모듈 import도 허용
    _INTERNAL_OK = {"numpy", "pandas", "openpyxl", "dateutil", "pytz"}

    def _safe_import(name, globals=None, locals=None, fromlist=(), level=0):
        root = name.split(".")[0]
        if root in _ALLOWED_MODULES or root in _INTERNAL_OK:
            return importlib.import_module(name)
        raise ImportError(f"'{name}' 모듈은 사용할 수 없습니다.")
    return _safe_import


def _safe_builtins() -> dict:
    import builtins
    safe = {
        name: getattr(builtins, name)
        for name in dir(builtins)
        if name not in _BLOCKED_BUILTINS and not name.startswith("_")
    }
    safe["__build_class__"] = builtins.__build_class__
    safe["__name__"] = "__main__"
    safe["__import__"] = _make_safe_import()
    return safe


def _make_save(saved: list[str], ns: dict):
    def save(filename: str, df: pd.DataFrame | None = None):
        df = df if df is not None else ns.get("result")
        if df is None:
            raise ValueError("`result` 변수가 없습니다. `result = ...` 로 지정하세요.")
        if not isinstance(df, pd.DataFrame):
            raise TypeError(f"DataFrame이 필요합니다. (현재: {type(df).__name__})")
        name = Path(filename).name
        if any(c in filename for c in ["/", "\\", ".."]):
            raise ValueError("파일명에 경로 문자를 사용할 수 없습니다.")
        suffix = Path(name).suffix.lower()
        stem = Path(name).stem
        # 파일명이 너무 generic하면 타임스탬프 추가
        if stem.lower() in ("output", "result", "결과", "출력"):
            ts = datetime.now().strftime("%m%d_%H%M%S")
            name = f"{stem}_{ts}{suffix}"
        if suffix not in (".xlsx", ".csv"):
            raise ValueError(f"지원하지 않는 형식: {suffix} (.xlsx 또는 .csv만 가능)")
        # 저장 전 정수형 정리
        df = df.copy()
        for col in df.select_dtypes(include="float").columns:
            non_null = df[col].dropna()
            if len(non_null) > 0 and (non_null % 1 == 0).all():
                try:
                    df[col] = df[col].astype("Int64")
                except Exception:
                    pass

        dest = RESULT_DIR / name
        if suffix == ".csv":
            df.to_csv(dest, index=False)
        else:
            import openpyxl
            df.to_excel(dest, index=False)
            wb = openpyxl.load_workbook(dest)
            ws = wb.active
            num_cols = {
                i + 1 for i, col in enumerate(df.columns)
                if pd.api.types.is_numeric_dtype(df[col])
            }
            text_cols = {
                i + 1 for i, col in enumerate(df.columns)
                if not pd.api.types.is_numeric_dtype(df[col])
            }

            def _cell_width(value) -> int:
                """한글(2) + 영문/숫자(1) 기준 표시 너비 계산."""
                if value is None:
                    return 0
                w = 0
                for ch in str(value):
                    # Hangul Syllables / Jamo / CJK 등
                    cp = ord(ch)
                    if (0xAC00 <= cp <= 0xD7AF or 0x1100 <= cp <= 0x11FF or
                            0x3130 <= cp <= 0x318F or 0x3400 <= cp <= 0x9FFF):
                        w += 2
                    else:
                        w += 1
                return w

            # 천단위 콤마 + 컬럼 너비 자동 조정
            col_widths: dict[int, int] = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.column in num_cols and cell.row > 1:
                        cell.number_format = "#,##0"
                    col_widths[cell.column] = max(
                        col_widths.get(cell.column, 0),
                        _cell_width(cell.value),
                    )
            for col_idx, width in col_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max(width + 2, 8), 60)

            # 텍스트 컬럼에서 연속된 같은 값 → 셀 병합 (원본 양식 복원)
            for col_idx in text_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                merge_start = 2
                prev_val = ws.cell(row=2, column=col_idx).value
                for row_idx in range(3, len(df) + 3):
                    cur_val = ws.cell(row=row_idx, column=col_idx).value
                    if cur_val != prev_val:
                        if row_idx - merge_start > 1:
                            ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{row_idx-1}")
                            ws.cell(row=merge_start, column=col_idx).alignment = \
                                openpyxl.styles.Alignment(vertical="center", wrap_text=True)
                        merge_start = row_idx
                        prev_val = cur_val
                # 마지막 그룹 처리
                last_row = len(df) + 1
                if last_row - merge_start > 0:
                    ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{last_row}")
                    ws.cell(row=merge_start, column=col_idx).alignment = \
                        openpyxl.styles.Alignment(vertical="center", wrap_text=True)

            wb.save(dest)
        saved.append(name)

    return save


class _PandasProxy:
    """pd.read_excel / pd.read_csv 호출을 이미 로드된 files 딕셔너리로 리다이렉트."""

    def __init__(self, files: dict):
        self._files = files

    def __getattr__(self, name):
        return getattr(pd, name)

    def _resolve(self, io) -> pd.DataFrame | None:
        fname = Path(str(io)).name
        if fname in self._files:
            return self._files[fname].copy()
        # 대소문자 무시 매칭
        fname_lower = fname.lower()
        for name, df in self._files.items():
            if name.lower() == fname_lower:
                return df.copy()
        return None

    def read_excel(self, io, *args, **kwargs):
        df = self._resolve(io)
        if df is not None:
            return df
        keys = list(self._files.keys())
        raise FileNotFoundError(
            f"'{Path(str(io)).name}' 파일 없음.\n"
            f"업로드된 파일: {keys}\n"
            f"→ files['{keys[0] if keys else '파일명'}'] 으로 접근하세요."
        )

    def read_csv(self, io, *args, **kwargs):
        df = self._resolve(io)
        if df is not None:
            return df
        keys = list(self._files.keys())
        raise FileNotFoundError(
            f"'{Path(str(io)).name}' 파일 없음.\n"
            f"업로드된 파일: {keys}\n"
            f"→ files['{keys[0] if keys else '파일명'}'] 으로 접근하세요."
        )


def _timeout_handler(signum, frame):
    raise TimeoutError("코드 실행 시간 초과 (30초)")


def execute(code: str, timeout: int = 30) -> ExecutionResult:
    code = _strip_provided_imports(code)

    issues = _validate(code)
    if issues:
        return ExecutionResult(
            success=False,
            error="코드 검증 실패:\n" + "\n".join(f"  - {i}" for i in issues),
        )

    saved: list[str] = []
    files = {f: read_file(f) for f in list_files() if read_file(f) is not None}

    ns: dict = {
        "files": files,
        "pd": _PandasProxy(files),
        "np": np,
        "result": None,
        "__builtins__": _safe_builtins(),
    }
    stdout = io.StringIO()
    ns["print"] = lambda *a, **kw: print(*a, **kw, file=stdout)
    ns["save"] = _make_save(saved, ns)

    in_main = threading.current_thread() is threading.main_thread()

    try:
        if in_main:
            signal.signal(signal.SIGALRM, _timeout_handler)
            signal.alarm(timeout)
        exec(compile(code, "<llm_generated>", "exec"), ns)  # noqa: S102
        if in_main:
            signal.alarm(0)
    except TimeoutError as e:
        return ExecutionResult(success=False, output=stdout.getvalue(), error=str(e))
    except Exception:
        if in_main:
            signal.alarm(0)
        error_msg = traceback.format_exc()
        # 실제 컬럼명을 에러 메시지에 추가해 AI가 다음 시도에서 참고할 수 있게
        if files:
            col_lines = "\n".join(
                f"  [{fname}] {list(df.columns)}" for fname, df in files.items()
            )
            error_msg += f"\n\n📋 실제 컬럼명 (참고):\n{col_lines}"
        return ExecutionResult(
            success=False, output=stdout.getvalue(), error=error_msg
        )

    result_df = ns.get("result")
    if not isinstance(result_df, pd.DataFrame):
        result_df = None
    else:
        # float 컬럼 중 소수점 없는 것(121.0 등) → 정수 변환
        for col in result_df.select_dtypes(include="float").columns:
            non_null = result_df[col].dropna()
            if len(non_null) > 0 and (non_null % 1 == 0).all():
                try:
                    result_df[col] = result_df[col].astype("Int64")
                except Exception:
                    pass

    return ExecutionResult(
        success=True,
        output=stdout.getvalue(),
        result_df=result_df,
        saved_files=saved,
    )
